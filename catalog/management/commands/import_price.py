"""Импорт каталога из прайса ЛЕСПРОФБАЗА (Сосна/ель + Лиственница + OSB).

Полностью пересобирает товары. Особые случаи:
 • Наличник «A × B × 2.2 м (3 м)» с ценой «300 (350)» → две карточки (2.2 м и 3 м).
 • Мебельный щит «18 × 200, 300, 400, 500, 600 × 3» → отдельные карточки по каждой ширине.
 • Планкен прямой/скошенный — обе породы в одной подкатегории (фильтр по породе).
 • Вагонка из липы и OSB-строки как надо. Сорта лиственницы — таблицей в описании.
"""
import re
from pathlib import Path
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction
import openpyxl

UNIT = {'м²': 'm2', 'м³': 'm3', 'пог. м': 'rm', 'пог.м': 'rm', 'шт': 'pc'}
GROUPS = ["Брус и брусок", "Вагонка", "Доска", "Мебельный щит",
          "Плиты", "Погонаж", "Фасадные материалы", "Элементы лестниц"]


def money(s):
    if s is None:
        return None, None
    s = str(s).replace('\xa0', ' ')
    unit = None
    for k, v in UNIT.items():
        if k in s:
            unit = v
            break
    m = re.search(r'(\d[\d\s]*)', s)
    return (int(m.group(1).replace(' ', '')) if m else None), unit


def dims3(text):
    """'16 × 120 × 6' → (16,120,6000). Длина (3-е число ≤20) в мм."""
    parts = re.split(r'[×хx]', str(text))
    nums = []
    for p in parts:
        m = re.search(r'\d+(?:[.,]\d+)?', p)
        nums.append(float(m.group(0).replace(',', '.')) if m else None)
    t = w = l = None
    if len(parts) >= 3 and nums[0] and nums[1] and nums[2] is not None:
        t, w = nums[0], nums[1]
        l = nums[2] * 1000 if nums[2] <= 20 else nums[2]
    elif len(parts) == 2 and nums[0] and nums[1] is not None:   # раскладка: ширина × длина
        w = nums[0]
        l = nums[1] * 1000 if nums[1] <= 20 else nums[1]
    return (int(t) if t else None, int(w) if w else None, int(l) if l else None)


def add(out, base, group, size, price, unit, species="pine_spruce",
        standard="", grade="", t=None, w=None, l=None, size_text=None, desc=""):
    out.append(dict(base=base, group=group, size=size, price=price, unit=unit,
                    species=species, standard=standard, grade=grade,
                    thickness=t, width=w, length=l,
                    size_text=size_text or size, desc=desc))


# ---------------- ХВОЯ ----------------
def parse_softwood(ws):
    sec_re = re.compile(r'^\s*\d+\.\s+([А-ЯЁ].*)')
    section = subgroup = None
    cur_price = cur_unit = None
    out = []
    for r in ws.iter_rows(values_only=True):
        F = str(r[5]).strip() if r[5] is not None else ''
        G = str(r[6]).strip() if r[6] is not None else ''
        J = r[9]
        if not F and not G:
            continue
        m = sec_re.match(F)
        if m:
            section = m.group(1).strip().upper(); subgroup = None
            cur_price = cur_unit = None
            continue
        if F and not re.match(r'^\d+$', F):
            pr, un = money(F)
            if pr and '₽' in F:
                cur_price, cur_unit = pr, un
            subgroup = F
            continue
        if not (re.match(r'^\d+$', F) and G):
            continue
        pr, un = money(J)
        if pr:
            cur_price, cur_unit = pr, un
        sub = (subgroup or "").lower()
        p, u = cur_price, cur_unit
        t, w, l = dims3(G)

        if section == "ВАГОНКА (ЛИПА)":
            continue
        elif section == "БРУСОК":
            base = "Рейка" if "рейка" in sub else "Брусок строганный"
            add(out, base, "Брус и брусок", G, p, u, t=t, w=w, l=l)
        elif section == "ПЛИНТУС":
            add(out, "Европлинтус" if "европлинтус" in sub else "Плинтус",
                "Погонаж", G, p, u, t=t, w=w, l=l)
        elif section == "УГОЛОК":
            add(out, "Уголок", "Погонаж", G, p, u, t=t, w=w, l=l)
        elif section == "ШТАПИК":
            add(out, "Штапик", "Погонаж", G, p, u, t=t, w=w, l=l)
        elif section == "РАСКЛАДКА":
            add(out, "Раскладка", "Погонаж", G, p, u, t=t, w=w, l=l)
        elif section == "НАЛИЧНИК":
            # 'A × B × 2.2 м (3 м)' цена 'X (Y)' → 2 карточки
            p2 = re.findall(r'\d[\d\s]*', str(J).replace('\xa0', ' '))
            price1 = int(p2[0].replace(' ', '')) if p2 else p
            price2 = int(p2[1].replace(' ', '')) if len(p2) > 1 else price1
            dm = re.findall(r'\d+', G.split('×')[0]) if '×' in G else []
            th = int(re.findall(r'\d+', G)[0]); wd = int(re.findall(r'\d+', G)[1])
            add(out, "Наличник", "Погонаж", f"{th} × {wd} × 2.2 м", price1, "pc",
                t=th, w=wd, l=2200)
            add(out, "Наличник", "Погонаж", f"{th} × {wd} × 3 м", price2, "pc",
                t=th, w=wd, l=3000)
        elif section == "МЕБЕЛЬНЫЙ ЩИТ":
            if "клееный брус" in sub:
                add(out, "Брус клееный", "Брус и брусок", G, p, u, grade="Сучковый", t=t, w=w, l=l)
            elif "тетива" in sub:
                add(out, "Тетива", "Элементы лестниц", G, p, u, grade="Сучковый", t=t, w=w, l=l)
            elif "перила" in sub:
                add(out, "Перила", "Элементы лестниц", G, p, u, grade="Без сучковый", size_text=G)
            elif "столб" in sub:
                add(out, "Столб", "Элементы лестниц", G, p, u, grade="Без сучковый", t=t, w=w, l=l)
            elif "балясина" in sub:
                add(out, "Балясина", "Элементы лестниц", G, p, u, grade="Без сучковый", t=t, w=w, l=l)
            else:
                # Сучковые: '18 × 200, 300, 400, 500, 600 × 3' → по каждой ширине
                mm = re.match(r'\s*(\d+)\s*×\s*([\d,\s]+?)\s*×\s*(\d+)', G)
                if mm:
                    th = int(mm.group(1)); ln = int(mm.group(3))
                    widths = [int(x) for x in re.findall(r'\d+', mm.group(2))]
                    for wd in widths:
                        add(out, "Мебельный щит", "Мебельный щит",
                            f"{th} × {wd} × {ln}", p, u, grade="Сучковый",
                            t=th, w=wd, l=ln * 1000)
                else:
                    add(out, "Мебельный щит", "Мебельный щит", G, p, u,
                        grade="Сучковый", t=t, w=w, l=l)
        elif section == "ВАГОНКА ШТИЛЬ":
            add(out, "Евровагонка" if "евровагонка" in sub else "Вагонка штиль",
                "Вагонка", G, p, u, t=t, w=w, l=l)
        elif section == "ИМИТАЦИЯ БРУСА":
            add(out, "Имитация бруса", "Фасадные материалы", G, p, u, t=t, w=w, l=l)
        elif section == "ПОЛОВАЯ ДОСКА":
            add(out, "Половая доска", "Доска", G, p, u, t=t, w=w, l=l)
        elif section == "ДОСКА СТРОГАННАЯ":
            add(out, "Доска строганная", "Доска", G, p, u, t=t, w=w, l=l)
        elif section == "БЛОК-ХАУС":
            add(out, "Блок-хаус", "Фасадные материалы", G, p, u, t=t, w=w, l=l)
        elif section == "ПЛАНКЕН ПРЯМОЙ":
            if "скошен" in sub or "косой" in sub:
                add(out, "Планкен скошенный (косой)", "Фасадные материалы", G, p, u, t=t, w=w, l=l)
            else:
                add(out, "Планкен прямой", "Фасадные материалы", G, p, u, t=t, w=w, l=l)
        elif section and section.startswith("ДОСКА ОБРЕЗНАЯ"):
            std = ""; grade = ""
            if "камерн" in sub: std = "Камерная сушка"
            elif "гост" in sub: std = "ГОСТ"
            elif "ту" in sub: std = "ТУ"
            if "второй сорт" in sub: grade = "2-й сорт"
            add(out, "Доска обрезная", "Доска", G, p, u, standard=std, grade=grade, t=t, w=w, l=l)
    return out


# ---------------- ЛИСТВЕННИЦА ----------------
def num(x):
    m = re.search(r'\d[\d\s]*', str(x))
    return int(m.group().replace(' ', '')) if m else None


def sort_table(sorts):
    rows = "".join(f"<tr><td>{k}</td><td>{num(v):,} ₽</td></tr>".replace(",", " ")
                   for k, v in sorts.items() if num(v))
    return ('<p class="sortnote">Цена зависит от сорта древесины (за м²):</p>'
            '<table class="sorts"><tr><th>Сорт</th><th>Цена</th></tr>' + rows + '</table>')


LARCH_MATRIX = {
    "вагонка из лиственницы": [("Вагонка лиственница", "Вагонка", "larch")],
    "террасная": [("Террасная доска лиственница", "Фасадные материалы", "larch")],
    "половая доска из лиственницы": [("Половая доска лиственница", "Доска", "larch")],
    "планкен прямой": [("Планкен прямой", "Фасадные материалы", "larch"),
                       ("Планкен скошенный (косой)", "Фасадные материалы", "larch")],
    "имитация бруса лиственница": [("Имитация бруса лиственница", "Фасадные материалы", "larch")],
    "вагонка из ангарской сосны": [("Вагонка ангарская сосна", "Вагонка", "pine_spruce")],
}


def parse_larch(ws):
    CYR2LAT = {'А': 'A', 'В': 'B', 'С': 'C', 'Д': 'D'}

    def norm_sort(label):
        label = re.sub(r'сорт', '', label, flags=re.I).strip()
        if not label:
            return ''
        if label in ('Экстра', 'Прима'):
            return label
        return ''.join(CYR2LAT.get(ch, ch) for ch in label)

    out = []
    targets = None; mode = None; cur_price = None; cur_sorts = None
    for r in ws.iter_rows(values_only=True):
        c = [(str(x).strip() if x is not None else '') for x in r]
        C, D, E, F, G = c[2], c[3], c[4], c[5], c[6]
        low = C.lower()
        matched = False
        for key, val in LARCH_MATRIX.items():
            if key in low:
                targets = val; mode = 'matrix'; cur_sorts = None; matched = True; break
        if matched:
            continue
        if "мебельный щит из лиственницы" in low:
            mode = 'shchit'; continue
        if "обрезная доска" in low:
            mode = 'obrezn'; cur_price = None; continue
        if "клееный брус из" in low:
            mode = 'glue'; cur_price = None; continue
        # строка-заголовок с сортами: запоминаем метки колонок
        if mode == 'matrix' and 'сорт' in (E + F + G).lower():
            cur_sorts = [norm_sort(x) for x in (c[4], c[5], c[6], c[7], c[8])]
            continue
        if not re.match(r'^\d', C) or not D:
            continue
        th = int(re.search(r'\d+', C).group())
        wm = re.search(r'\d+', D); wd = int(wm.group()) if wm else None

        if mode == 'matrix':
            sorts = cur_sorts or ['Экстра', 'Прима', 'AB', 'BC', 'CD']
            cols = [c[4], c[5], c[6], c[7], c[8]]
            for sortlbl, cell in zip(sorts, cols):
                pr = num(cell)
                if not pr or not sortlbl:
                    continue
                for base, group, sp in targets:
                    add(out, base, group, f"{th} × {wd} × 2–4 м", pr, "m2",
                        species=sp, grade=sortlbl, t=th, w=wd)
                    out[-1]['sort_in_name'] = True
        elif mode == 'obrezn':
            pr, _ = money(F)
            if pr:
                cur_price = pr
            ln = re.search(r'\d+', E); lm = int(ln.group()) * 1000 if ln else None
            add(out, "Доска обрезная лиственница", "Доска",
                f"{th} × {wd} × {E}", cur_price, "m3", species="larch", t=th, w=wd, l=lm)
        elif mode == 'glue':
            pr, _ = money(E)
            if pr:
                cur_price = pr
            add(out, "Брус клееный из лиственницы", "Брус и брусок",
                f"{th} × {wd} × 6", cur_price, "m3", species="larch", t=th, w=wd, l=6000)
        elif mode == 'shchit':
            price = num(G)
            if not price:
                continue
            grade = F or "Сращенная"
            widths = [int(x) for x in re.findall(r'\d+', D)]
            if not widths:
                widths = [wd]
            for wdi in widths:
                add(out, "Мебельный щит лиственница", "Мебельный щит",
                    f"{th} × {wdi}", price, "m2", species="larch", grade=grade, t=th, w=wdi)
    return out


def parse_osb(ws):
    out = []
    for r in ws.iter_rows(values_only=True):
        cells = [(str(x).strip() if x is not None else '') for x in r]
        th = size = price = None
        for x in cells:
            xx = x.replace(' ', '')
            if th is None and re.fullmatch(r'\d{1,2}', xx):
                th = int(xx); continue
            if size is None and re.search(r'[x×х]', x) and re.search(r'\d', x):
                size = x; continue
            if price is None and re.fullmatch(r'\d{3,6}', xx):
                price = int(xx)
        if th and size and price:
            sh = size.replace('x', '×').replace('X', '×').replace('х', '×').strip()
            add(out, "ОСП (OSB-3) плита", "Плиты", f"{th} мм", price, "sheet",
                species="other", t=th, size_text=f"лист {sh} м",
                desc=("Ориентированно-стружечная плита OSB-3 — прочная и влагостойкая. "
                      f"Для чернового пола, обшивки стен, кровли и опалубки. Лист {sh} м, толщина {th} мм."))
    return out


class Command(BaseCommand):
    help = "Импорт каталога из прайса"

    def add_arguments(self, parser):
        parser.add_argument("path", nargs="?", default=None)
        parser.add_argument("--dry", action="store_true")

    def handle(self, *args, **opts):
        path = opts["path"] or str(Path(settings.BASE_DIR) / "data" / "price.xlsx")
        wb = openpyxl.load_workbook(path, data_only=True)
        soft = parse_softwood(wb["Сосна и ель"])
        larch = parse_larch(wb["Лиственница"])
        osb_name = next((n for n in wb.sheetnames if "osb" in n.lower()), None)
        osb = parse_osb(wb[osb_name]) if osb_name else []
        items = [it for it in (soft + larch + osb) if it["price"]]

        # Сливаем лиственничные подкатегории в базовые (порода различается фильтром)
        RENAME = {
            "Брус клееный из лиственницы": "Брус клееный",
            "Вагонка лиственница": "Вагонка",
            "Доска обрезная лиственница": "Доска обрезная",
            "Половая доска лиственница": "Половая доска",
            "Мебельный щит лиственница": "Мебельный щит",
            "Имитация бруса лиственница": "Имитация бруса",
            "Террасная доска лиственница": "Террасная доска",
        }
        for it in items:
            it["desc_key"] = it["base"]          # описание берём по исходному имени (по породе)
            if it["base"] in RENAME:
                it["base"] = RENAME[it["base"]]

        for it in items:
            name = f"{it['base']} {it['size']}"
            if it.get("standard"):
                name += f", {it['standard']}"
            if it.get("sort_in_name") and it.get("grade"):
                name += f", сорт {it['grade']}"
            it["name"] = name

        from collections import Counter
        by_g = Counter(it["group"] for it in items)
        self.stdout.write(f"Всего товаров: {len(items)}")
        for g in GROUPS:
            self.stdout.write(f"  {by_g.get(g, 0):3d}  {g}")
        self.stdout.write("Породы: " + str(dict(Counter(it["species"] for it in items))))
        subcats = sorted(set((it["base"], it["group"]) for it in items))
        self.stdout.write(f"Подкатегорий: {len(subcats)}")
        if opts["dry"]:
            for b, g in subcats:
                self.stdout.write(f"    {g} / {b}")
            return

        from catalog.models import Category, Product
        try:
            from catalog.media_map import DESC_MAP
        except Exception:
            DESC_MAP = {}
        import html as _html
        esc = _html.escape
        NOTE = ('<div class="desc-note"><h4>Условия и сервис</h4><ul class="ticks">'
                '<li>Доставка по Москве и Московской области — круглосуточно</li>'
                '<li>Распил и подбор под ваш размер и задачу</li>'
                '<li>Проверяем каждую партию перед отгрузкой</li>'
                '<li>Для частных лиц — оплата после доставки</li>'
                '</ul></div>')

        def list_html(block):
            lis = []
            for ln in [x.strip() for x in block.split("\n") if x.strip()]:
                done = False
                for dash in (" — ", " – ", " - "):
                    if dash in ln:
                        a, _, b = ln.partition(dash)
                        lis.append("<li><b>%s</b> — %s</li>" % (esc(a.strip()), esc(b.strip())))
                        done = True
                        break
                if not done:
                    lis.append("<li>%s</li>" % esc(ln))
            return '<ul class="ticks">' + "".join(lis) + "</ul>"

        def to_html(t):
            t = (t or "").strip()
            if not t:
                return ""
            blocks = [b.strip() for b in t.split("\n\n") if b.strip()]
            out, i = [], 0
            while i < len(blocks):
                b = blocks[i]
                low = b.lower()
                if "\n" not in b and b.endswith(":") and len(b) < 40:
                    out.append('<h4 class="desc-h">%s</h4>' % esc(b.rstrip(":")))
                    if i + 1 < len(blocks):
                        out.append(list_html(blocks[i + 1]))
                        i += 2
                        continue
                    i += 1
                    continue
                if low.startswith("применение") or low.startswith("область применения"):
                    head, _, rest = b.partition(":")
                    rest = rest.strip()
                    if rest:
                        rest = rest[:1].upper() + rest[1:]
                        out.append('<p class="desc-lead"><b>%s.</b> %s</p>' % (esc(head.strip()), esc(rest)))
                    else:
                        out.append('<h4 class="desc-h">%s</h4>' % esc(head.strip()))
                    i += 1
                    continue
                if "\n" in b:
                    out.append(list_html(b))
                    i += 1
                    continue
                out.append("<p>%s</p>" % esc(b))
                i += 1
            return "".join(out)
        with transaction.atomic():
            groups = {}
            for i, g in enumerate(GROUPS):
                groups[g], _ = Category.objects.get_or_create(name=g, parent=None, defaults={"order": i})
            Product.objects.all().delete()
            Category.objects.filter(parent__isnull=False).delete()
            sc = {}

            def subcat(base, group):
                if (base, group) not in sc:
                    sc[(base, group)], _ = Category.objects.get_or_create(name=base, parent=groups[group])
                return sc[(base, group)]
            n = 0
            for it in items:
                Product.objects.create(
                    category=subcat(it["base"], it["group"]), name=it["name"],
                    species=it["species"], grade=it.get("grade", ""), standard=it.get("standard", ""),
                    thickness=it.get("thickness"), width=it.get("width"), length=it.get("length"),
                    size_text=it["size_text"], unit=it["unit"], price=it["price"],
                    description=(to_html(DESC_MAP.get(it.get("desc_key", it["base"]), ""))
                                 + (it["desc"] if "sorts" in it.get("desc", "") else "")
                                 + NOTE),
                    in_stock=True, is_active=True)
                n += 1
            self.stdout.write(self.style.SUCCESS(f"Создано товаров: {n}, подкатегорий: {len(sc)}"))
